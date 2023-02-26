import csv
import json
import os

import openpyxl

# Task 1
encoded = b'r\xc3\xa9sum\xc3\xa9'
decoded = encoded.decode()
print(decoded)

encoded_2 = decoded.encode("Latin1")
print(encoded_2)

decoded_2 = encoded_2.decode("Latin1")
print(decoded_2)

# Task 2
s_1 = input("Введите 1-ю строку: ")
s_2 = input("Введите 2-ю строку: ")
s_3 = input("Введите 3-ю строку: ")
s_4 = input("Введите 4-ю строку: ")

f = open("/homeworks/task_6/files/file.txt", "w")
f.write(f"{s_1}{os.linesep}")
f.write(f"{s_2}{os.linesep}")
f.close()

f = open("/homeworks/task_6/files/file.txt", "r")
print(f.read())
f.close()

with open("/homeworks/task_6/files/file.txt", "a+") as f:
    f.write(f"{s_3}{os.linesep}")
    f.write(f"{s_4}{os.linesep}")
    f.seek(0)
    print(f.read())

# Task 3
d = {
    123456: ("Vitali", 25),
    234567: ("Yuri", 23),
    345678: ("Ann", 13),
    456789: ("Leo", 1),
    567890: ("Ina", 25)
}

with open("/homeworks/task_6/files/file.json", "w+") as f:
    json.dump(d, f)
    f.seek(0)
    print(json.load(f))

# Task 4
with open("/homeworks/task_6/files/file.json", "r") as f:
    data = json.load(f)

with open("/homeworks/task_6/files/file.csv", "w") as f:
    writer = csv.writer(f)

    headers = ("Id", "Name", "Age", "Phone")
    phones = ("+222", "+333", "+444", "+555", "+666")

    writer.writerow(headers)

    for key, value in data.items():
        writer.writerow(
            [key] + value
        )

with open("/homeworks/task_6/files/file.csv", "r") as f:
    reader = csv.reader(f)
    rows = [row for row in reader]

for i in range(1, len(rows)):
    rows[i].append(phones[i - 1])

with open("/homeworks/task_6/files/file.csv", "w") as f:
    writer = csv.writer(f)
    writer.writerows(rows)

# Task 5
wb = openpyxl.Workbook()
sheet = wb.active

with open("/homeworks/task_6/files/file.csv", "r") as f:
    reader = csv.reader(f)

    list_csv = list(reader)

    for i in range(len(list_csv)):
        if i > 0:
            sheet.cell(row=1, column=i + 1).value = f"Person {i}"

        for y in range(len(list_csv[i])):
            sheet.cell(row=y + 2, column=i + 1).value = list_csv[i][y]

wb.save("/Users/zaitsau/PycharmProjects/pythonProject/homeworks/task_6/files/file.xlsx")
wb.close()
